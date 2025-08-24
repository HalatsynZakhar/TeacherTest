import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def create_test_template(output_path: str):
    """
    Создает Excel шаблон для тестов с подробными инструкциями.
    
    Args:
        output_path: Путь для сохранения шаблона
    """
    # Создаем новую книгу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Шаблон теста"
    
    # Стили для заголовков и инструкций
    header_font = Font(bold=True, size=14, color="FFFFFF")
    instruction_font = Font(size=11, color="000000")
    example_font = Font(size=10, color="666666", italic=True)
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    instruction_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Инструкции в первой строке
    instructions = [
        "ИНСТРУКЦИЯ ПО ЗАПОЛНЕНИЮ ТЕСТА:",
        "1. Номер вопроса (колонка A): Укажите номер вопроса. Если несколько вопросов имеют одинаковый номер (например, 1, 1, 1, 2, 2), система выберет случайный вопрос из каждой группы.",
        "2. Вопрос (колонка B): Текст вопроса.",
        "3. Правильный ответ/Тип (колонка C): Для вопросов с вариантами ответов - буква правильного ответа (A, B, C, D). Для открытых вопросов - слово 'открытый'.",
        "4. Вес (колонка D): Числовое значение сложности вопроса (1-легкий, 5-сложный). Используется для сортировки от легкого к сложному.",
        "5. Варианты ответов (колонки E-H): Варианты A, B, C, D. Оставьте пустыми для открытых вопросов.",
        "ВАЖНО: Заполнение начинается со строки 2! Не удаляйте эту инструкцию."
    ]
    
    # Объединяем ячейки для инструкций
    ws.merge_cells('A1:H1')
    ws['A1'] = '\n'.join(instructions)
    ws['A1'].font = instruction_font
    ws['A1'].fill = instruction_fill
    ws['A1'].alignment = Alignment(wrap_text=True, vertical='top')
    ws['A1'].border = border
    
    # Устанавливаем высоту первой строки
    ws.row_dimensions[1].height = 150
    
    # Заголовки колонок во второй строке
    headers = ['Номер вопроса', 'Вопрос', 'Правильный ответ/Тип', 'Вес', 'Вариант A', 'Вариант B', 'Вариант C', 'Вариант D']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Примеры данных
    examples = [
        [1, "Какой язык программирования используется для веб-разработки?", "A", 2, "JavaScript", "Python", "C++", "Java"],
        [1, "Что такое HTML?", "B", 1, "Язык программирования", "Язык разметки", "База данных", "Операционная система"],
        [2, "Объясните принцип работы HTTP протокола", "открытый", 4, "", "", "", ""],
        [3, "Сколько будет 2+2?", "C", 1, "3", "5", "4", "6"]
    ]
    
    for row_idx, example in enumerate(examples, 3):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = example_font
            cell.border = border
            if col_idx <= 4:  # Основные колонки
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Настройка ширины колонок
    column_widths = [15, 50, 20, 10, 25, 25, 25, 25]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # Создаем директорию если не существует
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # Сохраняем файл
    wb.save(output_path)
    
def create_answer_key_template(output_path: str):
    """
    Создает Excel шаблон для ключа ответов с инструкциями.
    
    Args:
        output_path: Путь для сохранения шаблона
    """
    # Создаем новую книгу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ключ ответов"
    
    # Стили
    header_font = Font(bold=True, size=14, color="FFFFFF")
    instruction_font = Font(size=11, color="000000")
    example_font = Font(size=10, color="666666", italic=True)
    
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    instruction_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Инструкции в первой строке
    instructions = [
        "ИНСТРУКЦИЯ ПО ЗАПОЛНЕНИЮ КЛЮЧА ОТВЕТОВ:",
        "1. Номер вопроса (колонка A): Номер вопроса из теста.",
        "2. Правильный ответ (колонка B): Для вопросов с вариантами - буква (A, B, C, D). Для открытых вопросов - точный текст ответа.",
        "3. Вес (колонка C): Количество баллов за правильный ответ.",
        "ВАЖНО: Заполнение начинается со строки 2! Не удаляйте эту инструкцию."
    ]
    
    # Объединяем ячейки для инструкций
    ws.merge_cells('A1:C1')
    ws['A1'] = '\n'.join(instructions)
    ws['A1'].font = instruction_font
    ws['A1'].fill = instruction_fill
    ws['A1'].alignment = Alignment(wrap_text=True, vertical='top')
    ws['A1'].border = border
    
    # Устанавливаем высоту первой строки
    ws.row_dimensions[1].height = 100
    
    # Заголовки колонок
    headers = ['Номер вопроса', 'Правильный ответ', 'Вес']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Примеры данных
    examples = [
        [1, "A", 2],
        [2, "Протокол передачи гипертекста", 4],
        [3, "C", 1]
    ]
    
    for row_idx, example in enumerate(examples, 3):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = example_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Настройка ширины колонок
    column_widths = [20, 40, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # Создаем директорию если не существует
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # Сохраняем файл
    wb.save(output_path)