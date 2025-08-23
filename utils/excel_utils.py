"""
Утилиты для работы с Excel
"""
import os
import logging
from typing import List, Dict, Tuple, Any, Optional, Union
from pathlib import Path
import tempfile
import time
import io

import pandas as pd
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension
from openpyxl.worksheet.table import Table, TableStyleInfo

logger = logging.getLogger(__name__)

def open_workbook(file_path: str) -> Workbook:
    """
    Открывает Excel-файл и возвращает объект Workbook.
    Создает новый файл, если указанный не существует.
    
    Args:
        file_path (str): Путь к Excel-файлу
    
    Returns:
        Workbook: Объект рабочей книги
    
    Raises:
        OSError: Если возникла ошибка при открытии файла
    """
    try:
        if os.path.exists(file_path):
            logger.debug(f"Открытие существующего файла: {file_path}")
            return openpyxl.load_workbook(file_path)
        else:
            logger.info(f"Файл не найден, создаем новый: {file_path}")
            return openpyxl.Workbook()
    except Exception as e:
        logger.error(f"Ошибка при открытии файла {file_path}: {e}")
        raise OSError(f"Не удалось открыть файл {file_path}: {e}")

def save_workbook(workbook: Workbook, file_path: str) -> bool:
    """
    Сохраняет рабочую книгу Excel в указанный файл.
    
    Args:
        workbook (Workbook): Объект рабочей книги для сохранения
        file_path (str): Путь для сохранения файла
    
    Returns:
        bool: True, если сохранение успешно, False в случае ошибки
    """
    try:
        # Создаем директорию, если она не существует
        directory = os.path.dirname(file_path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory)
            
        workbook.save(file_path)
        logger.info(f"Файл успешно сохранен: {file_path}")
        return True
    except Exception as e:
        logger.error(f"Ошибка при сохранении файла {file_path}: {e}")
        return False

def get_cell_value(worksheet: Worksheet, row: int, column: Union[int, str]) -> Any:
    """
    Получает значение ячейки из заданной позиции.
    
    Args:
        worksheet (Worksheet): Рабочий лист
        row (int): Номер строки (от 1)
        column (Union[int, str]): Номер столбца (от 1) или буквенное обозначение
    
    Returns:
        Any: Значение ячейки
    """
    try:
        if isinstance(column, str):
            cell = worksheet.cell(row=row, column=column_index_from_string(column))
        else:
            cell = worksheet.cell(row=row, column=column)
        
        logger.debug(f"Получено значение ячейки [{row}, {column}]: {cell.value}")
        return cell.value
    except Exception as e:
        logger.error(f"Ошибка при получении значения ячейки [{row}, {column}]: {e}")
        return None

def set_cell_value(worksheet: Worksheet, row: int, column: Union[int, str], value: Any) -> bool:
    """
    Устанавливает значение ячейки в заданной позиции.
    
    Args:
        worksheet (Worksheet): Рабочий лист
        row (int): Номер строки (от 1)
        column (Union[int, str]): Номер столбца (от 1) или буквенное обозначение
        value (Any): Устанавливаемое значение
    
    Returns:
        bool: True, если значение успешно установлено
    """
    try:
        if isinstance(column, str):
            cell = worksheet.cell(row=row, column=column_index_from_string(column))
        else:
            cell = worksheet.cell(row=row, column=column)
        
        cell.value = value
        logger.debug(f"Установлено значение ячейки [{row}, {column}]: {value}")
        return True
    except Exception as e:
        logger.error(f"Ошибка при установке значения ячейки [{row}, {column}]: {e}")
        return False

def find_column_by_header(worksheet: Worksheet, header_text: str, header_row: int = 1) -> Optional[int]:
    """
    Находит номер столбца по тексту заголовка.
    
    Args:
        worksheet (Worksheet): Рабочий лист
        header_text (str): Текст для поиска в заголовках
        header_row (int, optional): Номер строки с заголовками. По умолчанию 1.
    
    Returns:
        Optional[int]: Номер столбца (от 1) или None, если не найден
    """
    try:
        for column in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=header_row, column=column).value
            if cell_value and str(cell_value).strip() == header_text.strip():
                logger.debug(f"Найден столбец '{header_text}' на позиции {column}")
                return column
        
        logger.warning(f"Столбец с заголовком '{header_text}' не найден")
        return None
    except Exception as e:
        logger.error(f"Ошибка при поиске столбца с заголовком '{header_text}': {e}")
        return None

def get_range_values(worksheet: Worksheet, start_row: int, start_column: Union[int, str], 
                    end_row: Optional[int] = None, end_column: Optional[Union[int, str]] = None) -> List[List[Any]]:
    """
    Получает значения из диапазона ячеек.
    
    Args:
        worksheet (Worksheet): Рабочий лист
        start_row (int): Начальная строка
        start_column (Union[int, str]): Начальный столбец (номер или буква)
        end_row (Optional[int], optional): Конечная строка. Если None, до конца данных.
        end_column (Optional[Union[int, str]], optional): Конечный столбец. Если None, до конца данных.
    
    Returns:
        List[List[Any]]: Список значений из диапазона
    """
    try:
        # Преобразуем буквенные обозначения столбцов в числа
        if isinstance(start_column, str):
            start_column = column_index_from_string(start_column)
        
        if isinstance(end_column, str):
            end_column = column_index_from_string(end_column)
        
        # Если конечные значения не указаны, используем максимальные
        if end_row is None:
            end_row = worksheet.max_row
        
        if end_column is None:
            end_column = worksheet.max_column
        
        # Получаем значения
        values = []
        for row in range(start_row, end_row + 1):
            row_values = []
            for col in range(start_column, end_column + 1):
                row_values.append(worksheet.cell(row=row, column=col).value)
            values.append(row_values)
        
        logger.debug(f"Получены значения из диапазона [{start_row}, {start_column}] - [{end_row}, {end_column}]")
        return values
    except Exception as e:
        logger.error(f"Ошибка при получении значений из диапазона: {e}")
        return []

def set_column_width(worksheet: Worksheet, column: Union[int, str], width: float) -> bool:
    """
    Устанавливает ширину столбца.
    
    Args:
        worksheet (Worksheet): Рабочий лист
        column (Union[int, str]): Номер столбца или буквенное обозначение
        width (float): Ширина столбца
    
    Returns:
        bool: True, если успешно
    """
    try:
        if isinstance(column, int):
            column_letter = get_column_letter(column)
        else:
            column_letter = column
        
        worksheet.column_dimensions[column_letter].width = width
        logger.debug(f"Установлена ширина столбца {column_letter}: {width}")
        return True
    except Exception as e:
        logger.error(f"Ошибка при установке ширины столбца {column}: {e}")
        return False

def set_row_height(worksheet: Worksheet, row: int, height: float) -> bool:
    """
    Устанавливает высоту строки.
    
    Args:
        worksheet (Worksheet): Рабочий лист
        row (int): Номер строки
        height (float): Высота строки
    
    Returns:
        bool: True, если успешно
    """
    try:
        worksheet.row_dimensions[row].height = height
        logger.debug(f"Установлена высота строки {row}: {height}")
        return True
    except Exception as e:
        logger.error(f"Ошибка при установке высоты строки {row}: {e}")
        return False

def apply_style_to_cell(worksheet: Worksheet, row: int, column: Union[int, str], 
                       bold: bool = False, font_size: int = 11, font_name: str = 'Calibri',
                       alignment: Dict = None, border: Dict = None, fill_color: str = None) -> bool:
    """
    Применяет стили к ячейке.
    
    Args:
        worksheet (Worksheet): Рабочий лист
        row (int): Номер строки
        column (Union[int, str]): Номер столбца или буквенное обозначение
        bold (bool, optional): Жирный шрифт. По умолчанию False.
        font_size (int, optional): Размер шрифта. По умолчанию 11.
        font_name (str, optional): Название шрифта. По умолчанию 'Calibri'.
        alignment (Dict, optional): Выравнивание {'horizontal': 'center', 'vertical': 'center', 'wrap_text': True}
        border (Dict, optional): Границы {'style': 'thin', 'color': '000000'}
        fill_color (str, optional): Цвет заливки в формате RRGGBB
    
    Returns:
        bool: True, если успешно
    """
    try:
        if isinstance(column, str):
            column = column_index_from_string(column)
        
        cell = worksheet.cell(row=row, column=column)
        
        # Применяем шрифт
        font = Font(name=font_name, size=font_size, bold=bold)
        cell.font = font
        
        # Применяем выравнивание
        if alignment:
            align_params = {}
            if 'horizontal' in alignment:
                align_params['horizontal'] = alignment['horizontal']
            if 'vertical' in alignment:
                align_params['vertical'] = alignment['vertical']
            if 'wrap_text' in alignment:
                align_params['wrap_text'] = alignment['wrap_text']
                
            cell.alignment = Alignment(**align_params)
        
        # Применяем границы
        if border:
            border_style = border.get('style', 'thin')
            border_color = border.get('color', '000000')
            side = Side(style=border_style, color=border_color)
            cell.border = Border(left=side, right=side, top=side, bottom=side)
        
        # Применяем заливку
        if fill_color:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        
        logger.debug(f"Применены стили к ячейке [{row}, {column}]")
        return True
    except Exception as e:
        logger.error(f"Ошибка при применении стилей к ячейке [{row}, {column}]: {e}")
        return False

def apply_style_to_range(worksheet: Worksheet, start_row: int, start_column: Union[int, str],
                        end_row: int, end_column: Union[int, str],
                        bold: bool = False, font_size: int = 11, font_name: str = 'Calibri',
                        alignment: Dict = None, border: Dict = None, fill_color: str = None) -> bool:
    """
    Применяет стили к диапазону ячеек.
    
    Args:
        worksheet (Worksheet): Рабочий лист
        start_row (int): Начальная строка
        start_column (Union[int, str]): Начальный столбец
        end_row (int): Конечная строка
        end_column (Union[int, str]): Конечный столбец
        bold (bool, optional): Жирный шрифт. По умолчанию False.
        font_size (int, optional): Размер шрифта. По умолчанию 11.
        font_name (str, optional): Название шрифта. По умолчанию 'Calibri'.
        alignment (Dict, optional): Выравнивание {'horizontal': 'center', 'vertical': 'center', 'wrap_text': True}
        border (Dict, optional): Границы {'style': 'thin', 'color': '000000'}
        fill_color (str, optional): Цвет заливки в формате RRGGBB
    
    Returns:
        bool: True, если успешно
    """
    try:
        # Преобразуем буквенные обозначения столбцов в числа
        if isinstance(start_column, str):
            start_column = column_index_from_string(start_column)
        
        if isinstance(end_column, str):
            end_column = column_index_from_string(end_column)
        
        # Применяем стили к каждой ячейке в диапазоне
        success = True
        for row in range(start_row, end_row + 1):
            for col in range(start_column, end_column + 1):
                cell_success = apply_style_to_cell(
                    worksheet, row, col, bold, font_size, font_name, 
                    alignment, border, fill_color
                )
                success = success and cell_success
        
        logger.debug(f"Применены стили к диапазону [{start_row}, {start_column}] - [{end_row}, {end_column}]")
        return success
    except Exception as e:
        logger.error(f"Ошибка при применении стилей к диапазону: {e}")
        return False

def create_table_from_data(worksheet: Worksheet, data: List[List[Any]], start_row: int, start_column: int,
                          table_name: str, table_style: str = 'TableStyleMedium2',
                          headers: List[str] = None) -> bool:
    """
    Создает таблицу из данных.
    
    Args:
        worksheet (Worksheet): Рабочий лист
        data (List[List[Any]]): Данные для таблицы (без заголовков)
        start_row (int): Начальная строка для размещения таблицы
        start_column (int): Начальный столбец для размещения таблицы
        table_name (str): Уникальное имя таблицы
        table_style (str, optional): Стиль таблицы. По умолчанию 'TableStyleMedium2'.
        headers (List[str], optional): Заголовки столбцов. Если не указаны, используется первая строка данных.
    
    Returns:
        bool: True, если успешно
    """
    try:
        if not data:
            logger.warning("Нет данных для создания таблицы")
            return False
        
        # Определяем размеры таблицы
        rows_count = len(data)
        cols_count = len(data[0]) if data else 0
        
        # Записываем заголовки, если они указаны
        current_row = start_row
        if headers:
            for col_idx, header in enumerate(headers, start=start_column):
                worksheet.cell(row=current_row, column=col_idx).value = header
            
            # Применяем стиль к заголовкам
            apply_style_to_range(
                worksheet, current_row, start_column, current_row, start_column + len(headers) - 1,
                bold=True, alignment={'horizontal': 'center', 'vertical': 'center'}
            )
            
            current_row += 1
        
        # Записываем данные
        for row_idx, row_data in enumerate(data, start=current_row):
            for col_idx, cell_value in enumerate(row_data, start=start_column):
                worksheet.cell(row=row_idx, column=col_idx).value = cell_value
        
        # Определяем диапазон таблицы
        end_row = start_row + rows_count - (0 if not headers else -1)
        end_column = start_column + cols_count - 1
        
        # Создаем и добавляем таблицу
        table_ref = f"{get_column_letter(start_column)}{start_row}:{get_column_letter(end_column)}{end_row}"
        table = Table(displayName=table_name, ref=table_ref)
        
        # Устанавливаем стиль таблицы
        style = TableStyleInfo(
            name=table_style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        
        # Добавляем таблицу на лист
        worksheet.add_table(table)
        logger.info(f"Создана таблица '{table_name}' в диапазоне {table_ref}")
        
        return True
    except Exception as e:
        logger.error(f"Ошибка при создании таблицы из данных: {e}")
        return False

def convert_column_letter_to_index(column_letter: str) -> int:
    """
    Преобразует буквенное обозначение столбца в числовой индекс.
    
    Args:
        column_letter (str): Буквенное обозначение столбца (A, B, C, ..., AA, AB, ...)
    
    Returns:
        int: Числовой индекс столбца (1, 2, 3, ...)
    """
    try:
        return column_index_from_string(column_letter)
    except Exception as e:
        logger.error(f"Ошибка при преобразовании букв столбца '{column_letter}' в индекс: {e}")
        return -1

def convert_column_index_to_letter(column_index: int) -> str:
    """
    Преобразует числовой индекс столбца в буквенное обозначение.
    
    Args:
        column_index (int): Числовой индекс столбца (1, 2, 3, ...)
    
    Returns:
        str: Буквенное обозначение столбца (A, B, C, ..., AA, AB, ...)
    """
    try:
        return get_column_letter(column_index)
    except Exception as e:
        logger.error(f"Ошибка при преобразовании индекса столбца '{column_index}' в буквы: {e}")
        return ""

def merge_cells(worksheet: Worksheet, start_row: int, start_column: Union[int, str],
               end_row: int, end_column: Union[int, str]) -> bool:
    """
    Объединяет ячейки в заданном диапазоне.
    
    Args:
        worksheet (Worksheet): Рабочий лист
        start_row (int): Начальная строка
        start_column (Union[int, str]): Начальный столбец
        end_row (int): Конечная строка
        end_column (Union[int, str]): Конечный столбец
    
    Returns:
        bool: True, если успешно
    """
    try:
        # Преобразуем индексы в буквы, если необходимо
        if isinstance(start_column, int):
            start_column = get_column_letter(start_column)
        
        if isinstance(end_column, int):
            end_column = get_column_letter(end_column)
        
        # Объединяем ячейки
        merge_range = f"{start_column}{start_row}:{end_column}{end_row}"
        worksheet.merge_cells(merge_range)
        
        logger.debug(f"Объединены ячейки в диапазоне {merge_range}")
        return True
    except Exception as e:
        logger.error(f"Ошибка при объединении ячеек: {e}")
        return False

def copy_worksheet(source_workbook: Workbook, source_worksheet_name: str, 
                 target_workbook: Workbook, target_worksheet_name: str = None) -> Optional[Worksheet]:
    """
    Копирует рабочий лист из одной книги в другую.
    
    Args:
        source_workbook (Workbook): Исходная рабочая книга
        source_worksheet_name (str): Имя исходного рабочего листа
        target_workbook (Workbook): Целевая рабочая книга
        target_worksheet_name (str, optional): Имя для нового рабочего листа. Если None, используется исходное имя.
    
    Returns:
        Optional[Worksheet]: Новый рабочий лист или None в случае ошибки
    """
    try:
        # Проверяем наличие исходного листа
        if source_worksheet_name not in source_workbook.sheetnames:
            logger.error(f"Исходный лист '{source_worksheet_name}' не найден")
            return None
        
        source_sheet = source_workbook[source_worksheet_name]
        
        # Определяем имя целевого листа
        if target_worksheet_name is None:
            target_worksheet_name = source_worksheet_name
        
        # Создаем новый лист
        target_sheet = target_workbook.create_sheet(title=target_worksheet_name)
        
        # Копируем данные и стили
        for row in source_sheet.rows:
            for cell in row:
                new_cell = target_sheet.cell(row=cell.row, column=cell.column)
                new_cell.value = cell.value
                
                # Копируем стиль
                if cell.has_style:
                    new_cell.font = cell.font
                    new_cell.border = cell.border
                    new_cell.fill = cell.fill
                    new_cell.number_format = cell.number_format
                    new_cell.protection = cell.protection
                    new_cell.alignment = cell.alignment
        
        # Копируем размеры строк и столбцов
        for col_letter, dimension in source_sheet.column_dimensions.items():
            target_sheet.column_dimensions[col_letter].width = dimension.width
            
        for row_num, dimension in source_sheet.row_dimensions.items():
            target_sheet.row_dimensions[row_num].height = dimension.height
        
        # Копируем объединенные ячейки
        for merged_cell_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_cell_range))
        
        logger.info(f"Лист '{source_worksheet_name}' успешно скопирован как '{target_worksheet_name}'")
        return target_sheet
    except Exception as e:
        logger.error(f"Ошибка при копировании рабочего листа: {e}")
        return None

def insert_image(worksheet: Worksheet, image_path: str, anchor_cell: str, 
                width: Optional[int] = None, height: Optional[int] = None, 
                preserve_aspect_ratio: bool = True, background_color: Optional[str] = None) -> bool:
    """
    Вставляет изображение в рабочий лист.
    
    Args:
        worksheet (Worksheet): Рабочий лист
        image_path (str): Путь к изображению
        anchor_cell (str): Ячейка привязки изображения (например, 'A1')
        width (Optional[int], optional): Ширина изображения в пикселях
        height (Optional[int], optional): Высота изображения в пикселях
        preserve_aspect_ratio (bool, optional): Сохранять пропорции изображения. По умолчанию True.
        background_color (Optional[str], optional): Цвет фона ячейки в формате RRGGBB. По умолчанию None.
    
    Returns:
        bool: True, если успешно
    """
    try:
        if not os.path.exists(image_path):
            logger.error(f"Изображение не найдено: {image_path}")
            return False
        
        # Устанавливаем цвет фона ячейки, если указан
        if background_color:
            set_cell_background(worksheet, anchor_cell, background_color)
            logger.debug(f"Установлен цвет фона {background_color} для ячейки {anchor_cell}")
            
        # Создаем объект изображения
        img = XLImage(image_path)
        
        # Получаем оригинальные размеры изображения
        original_width = img.width
        original_height = img.height
        
        # Устанавливаем размеры с сохранением пропорций, если требуется
        if width is not None and height is not None and preserve_aspect_ratio:
            # Если указаны оба размера, но нужно сохранить пропорции,
            # используем ширину как основу и пересчитываем высоту
            aspect_ratio = original_height / original_width if original_width > 0 else 1.0
            img.width = width
            img.height = int(width * aspect_ratio)
            logger.debug(f"Изображение масштабировано с сохранением пропорций: {img.width}x{img.height}")
        else:
            # Иначе устанавливаем размеры как указано
            if width is not None:
                img.width = width
            if height is not None:
                img.height = height
            
        # Вставляем изображение
        worksheet.add_image(img, anchor_cell)
        
        logger.debug(f"Изображение вставлено в ячейку {anchor_cell}, размеры: {img.width}x{img.height}")
        return True
    except Exception as e:
        logger.error(f"Ошибка при вставке изображения в ячейку {anchor_cell}: {e}")
        return False

def insert_image_from_buffer(worksheet: Worksheet, image_buffer, anchor_cell: str,
                           width: Optional[int] = None, height: Optional[int] = None,
                           preserve_aspect_ratio: bool = True, 
                           anchor_type: str = "oneCellAnchor",
                           background_color: Optional[str] = None) -> bool:
    """
    Вставляет изображение из буфера в рабочий лист.
    
    Args:
        worksheet (Worksheet): Рабочий лист
        image_buffer: Буфер с изображением (io.BytesIO)
        anchor_cell (str): Ячейка привязки изображения (например, 'A1')
        width (Optional[int], optional): Ширина изображения в пикселях
        height (Optional[int], optional): Высота изображения в пикселях
        preserve_aspect_ratio (bool, optional): Сохранять пропорции изображения. По умолчанию True.
        anchor_type (str, optional): Тип привязки изображения ('oneCellAnchor', 'absoluteAnchor', 'twoCellAnchor').
                                    По умолчанию 'oneCellAnchor'.
        background_color (Optional[str], optional): Цвет фона ячейки в формате RRGGBB. По умолчанию None.
    
    Returns:
        bool: True, если успешно
    """
    try:
        # Проверяем буфер на наличие данных
        buffer_size = image_buffer.getbuffer().nbytes
        if buffer_size == 0:
            logger.error(f"Пустой буфер изображения для ячейки {anchor_cell}")
            return False
            
        logger.debug(f"Вставка изображения из буфера в ячейку {anchor_cell}. Размер буфера: {buffer_size/1024:.2f} КБ")
        
        # Устанавливаем цвет фона ячейки, если указан
        if background_color:
            set_cell_background(worksheet, anchor_cell, background_color)
            logger.debug(f"Установлен цвет фона {background_color} для ячейки {anchor_cell}")
        
        # Сбрасываем указатель буфера в начало
        image_buffer.seek(0)
        
        # Получаем размеры изображения из буфера перед созданием объекта XLImage
        try:
            from PIL import Image as PILImage
            img = PILImage.open(image_buffer)
            original_width, original_height = img.size
            logger.debug(f"Размеры изображения из буфера: {original_width}x{original_height}")
            # Возвращаем указатель в начало буфера после чтения размеров
            image_buffer.seek(0)
        except Exception as e:
            logger.warning(f"Не удалось определить размеры изображения из буфера: {e}")
            original_width, original_height = 100, 100  # Значения по умолчанию
        
        # Создаем временный файл для надежной вставки изображения
        # Используем delete=False, чтобы файл не был удален автоматически
        with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as temp_file:
            temp_path = temp_file.name
            temp_file.write(image_buffer.getvalue())
            logger.debug(f"Создан временный файл для вставки: {temp_path}")
        
        try:
            # Создаем объект Image с временного файла
            img = openpyxl.drawing.image.Image(temp_path)
            
            # Устанавливаем размеры с учетом соотношения сторон
            if width is not None and height is not None:
                if preserve_aspect_ratio:
                    # Рассчитываем соотношение сторон
                    aspect_ratio = original_height / original_width if original_width > 0 else 1.0
                    img.width = width
                    img.height = int(width * aspect_ratio)
                    logger.debug(f"Установлены размеры с сохранением пропорций: {img.width}x{img.height} (соотношение: {aspect_ratio:.2f})")
                else:
                    img.width = width
                    img.height = height
                    logger.debug(f"Установлены фиксированные размеры: {img.width}x{img.height}")
            elif width is not None:
                img.width = width
                # Автоматически рассчитываем высоту для сохранения пропорций
                aspect_ratio = original_height / original_width if original_width > 0 else 1.0
                img.height = int(width * aspect_ratio)
                logger.debug(f"Установлена ширина {img.width} и автоматическая высота {img.height}")
            elif height is not None:
                img.height = height
                # Автоматически рассчитываем ширину для сохранения пропорций
                aspect_ratio = original_width / original_height if original_height > 0 else 1.0
                img.width = int(height * aspect_ratio)
                logger.debug(f"Установлена высота {img.height} и автоматическая ширина {img.width}")
            
            # Вставляем изображение в ячейку
            img.anchor = anchor_cell
            
            # Устанавливаем тип привязки изображения
            if hasattr(img, 'anchor_type') and anchor_type in ['oneCellAnchor', 'absoluteAnchor', 'twoCellAnchor']:
                img.anchor_type = anchor_type
                logger.debug(f"Установлен тип привязки: {anchor_type}")
            
            # Проверяем и при необходимости увеличиваем высоту строки
            row_num = int(''.join(filter(str.isdigit, anchor_cell)))
            if row_num > 0 and img.height is not None:
                # Преобразуем пиксели в единицы Excel (приблизительно)
                excel_height = img.height * 0.75  # коэффициент конвертации
                
                # Устанавливаем высоту строки, если она меньше необходимой
                current_height = worksheet.row_dimensions.get(row_num, 0)
                current_height = current_height.height if hasattr(current_height, 'height') else 0
                
                if current_height < excel_height:
                    worksheet.row_dimensions[row_num].height = excel_height
                    logger.debug(f"Увеличена высота строки {row_num} до {excel_height}")
            
            worksheet.add_image(img, anchor_cell)
            
            # ВАЖНО: НЕ удаляем временный файл, он нужен до сохранения книги
            # Сохраняем путь к файлу в глобальный список для последующей очистки
            if not hasattr(worksheet, '_temp_image_files'):
                worksheet._temp_image_files = []
            worksheet._temp_image_files.append(temp_path)
            logger.debug(f"Временный файл {temp_path} добавлен в список для последующей очистки")
            
            return True
            
        except Exception as add_e:
            logger.error(f"Ошибка при добавлении изображения в ячейку {anchor_cell}: {add_e}")
            try:
                os.unlink(temp_path)  # Удаляем только в случае ошибки
                logger.debug(f"Удален временный файл после ошибки: {temp_path}")
            except Exception as del_e:
                logger.warning(f"Не удалось удалить временный файл {temp_path}: {del_e}")
            return False
            
    except Exception as e:
        logger.error(f"Ошибка при вставке изображения из буфера в ячейку {anchor_cell}: {e}")
        return False

def auto_adjust_column_width(worksheet: Worksheet, columns: List[Union[int, str]] = None, 
                           min_width: float = 8, max_width: float = 50, 
                           padding: float = 1.5) -> bool:
    """
    Автоматически регулирует ширину столбцов на основе их содержимого.
    
    Args:
        worksheet (Worksheet): Рабочий лист
        columns (List[Union[int, str]], optional): Список столбцов для регулировки. Если None, все столбцы.
        min_width (float, optional): Минимальная ширина столбца. По умолчанию 8.
        max_width (float, optional): Максимальная ширина столбца. По умолчанию 50.
        padding (float, optional): Дополнительное пространство. По умолчанию 1.5.
    
    Returns:
        bool: True, если успешно
    """
    try:
        # Если столбцы не указаны, используем все столбцы
        if columns is None:
            columns = list(range(1, worksheet.max_column + 1))
        
        # Преобразуем буквенные обозначения в индексы
        column_indices = []
        for col in columns:
            if isinstance(col, str):
                column_indices.append(column_index_from_string(col))
            else:
                column_indices.append(col)
        
        # Для каждого столбца определяем максимальную ширину содержимого
        for col_idx in column_indices:
            col_letter = get_column_letter(col_idx)
            max_content_width = 0
            
            # Проверяем содержимое каждой ячейки в столбце
            for row_idx in range(1, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value:
                    # Оцениваем ширину по длине текста
                    # Это упрощенная оценка, может потребоваться доработка
                    content_width = len(str(cell.value)) * 1.2
                    max_content_width = max(max_content_width, content_width)
            
            # Применяем ограничения и добавляем отступ
            adjusted_width = min(max(min_width, max_content_width + padding), max_width)
            worksheet.column_dimensions[col_letter].width = adjusted_width
            
            logger.debug(f"Автоматически установлена ширина столбца {col_letter}: {adjusted_width}")
        
        return True
    except Exception as e:
        logger.error(f"Ошибка при автоматической регулировке ширины столбцов: {e}")
        return False

def clear_worksheet(worksheet: Worksheet, keep_first_row: bool = False) -> bool:
    """
    Очищает рабочий лист от данных.
    
    Args:
        worksheet (Worksheet): Рабочий лист для очистки
        keep_first_row (bool, optional): Сохранить первую строку (заголовки). По умолчанию False.
    
    Returns:
        bool: True, если успешно
    """
    try:
        start_row = 2 if keep_first_row else 1
        
        for row in range(start_row, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row, column=col).value = None
        
        logger.info(f"Рабочий лист очищен (с сохранением первой строки: {keep_first_row})")
        return True
    except Exception as e:
        logger.error(f"Ошибка при очистке рабочего листа: {e}")
        return False

def create_excel_copy(excel_file: str, output_dir: str) -> str:
    """
    Создает копию Excel файла в указанной директории
    
    Args:
        excel_file (str): Путь к исходному Excel файлу
        output_dir (str): Директория для создания копии
        
    Returns:
        str: Путь к созданной копии
    """
    import shutil
    
    try:
        if not os.path.exists(excel_file):
            logger.error(f"Исходный файл не найден: {excel_file}")
            raise FileNotFoundError(f"Исходный файл не найден: {excel_file}")
        
        if not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
            logger.info(f"Создана директория: {output_dir}")
        
        # Формируем имя для копии
        filename = os.path.basename(excel_file)
        temp_filename = f"temp_{filename}"
        output_path = os.path.join(output_dir, temp_filename)
        
        # Создаем копию файла
        shutil.copy2(excel_file, output_path)
        logger.info(f"Создана копия файла: {output_path}")
        
        return output_path
    except Exception as e:
        logger.error(f"Ошибка при создании копии файла: {e}")
        raise

def process_excel_file(excel_file: str, article_column: str, image_column: str, 
                      images_folder: str, start_row: int = 1, sheet_index: int = 1,
                      max_file_size_mb: int = 20, target_width: int = 300, target_height: int = 300,
                      adjust_cell_size: bool = False, column_width: int = 150,
                      row_height: int = 120) -> Dict[str, Any]:
    """
    Обрабатывает Excel файл, добавляя изображения по артикулам
    
    Args:
        excel_file (str): Путь к файлу Excel
        article_column (str): Столбец с артикулами
        image_column (str): Столбец для вставки изображений
        images_folder (str): Папка с изображениями
        start_row (int): Начальная строка (начиная с 1)
        sheet_index (int): Индекс листа в Excel-файле (начиная с 1)
        max_file_size_mb (int): Максимальный размер файла изображения в МБ
        target_width (int): Максимальная ширина изображения в пикселях (используется только для оптимизации)
        target_height (int): Максимальная высота изображения в пикселях (используется только для оптимизации)
        adjust_cell_size (bool): Изменять размер ячейки для лучшего отображения изображений
        column_width (int): Желаемая ширина колонки с изображениями в пикселях (при adjust_cell_size=True)
        row_height (int): Желаемая высота строки с изображениями в пикселях (при adjust_cell_size=True)
        
    Returns:
        Dict[str, Any]: Статистика обработки
    """
    import os
    import time
    import logging
    import tempfile
    from pathlib import Path
    from PIL import Image as PILImage
    from . import image_utils
    
    logger = logging.getLogger(__name__)
    
    try:
        logger.info(f"Начинаем обработку Excel файла: {excel_file}")
        logger.info(f"Параметры: article_column={article_column}, image_column={image_column}, start_row={start_row}, sheet_index={sheet_index}")
        logger.info(f"Папка с изображениями: {images_folder}")
        logger.info(f"Настройки изображений: max_file_size_mb={max_file_size_mb}")
        logger.info(f"Настройки форматирования: adjust_cell_size={adjust_cell_size}, column_width={column_width}, row_height={row_height}")
        
        # Конвертируем МБ в КБ для обработки
        max_size_kb = max_file_size_mb * 1024
        
        # Проверяем существование файла Excel
        if not os.path.exists(excel_file):
            logger.error(f"Файл Excel не найден: {excel_file}")
            raise FileNotFoundError(f"Файл Excel не найден: {excel_file}")
        
        # Проверяем существование папки с изображениями
        if not os.path.exists(images_folder):
            logger.error(f"Папка с изображениями не найдена: {images_folder}")
            raise FileNotFoundError(f"Папка с изображениями не найдена: {images_folder}")
        
        # Статистика
        stats = {
            "total_articles": 0,
            "images_found": 0,
            "images_inserted": 0,
            "start_time": time.time(),
            "end_time": None,
            "processing_time": None,
            "output_file": None
        }
        
        # Загружаем рабочую книгу
        logger.info(f"Загружаем рабочую книгу: {excel_file}")
        wb = openpyxl.load_workbook(excel_file)
        
        # Выбираем лист по индексу
        if sheet_index > len(wb.sheetnames):
            logger.error(f"Лист с индексом {sheet_index} не существует. Всего листов: {len(wb.sheetnames)}")
            raise ValueError(f"Лист с индексом {sheet_index} не существует. Всего листов: {len(wb.sheetnames)}")
        
        # Получаем имена всех листов
        logger.info(f"Листы в книге: {', '.join(wb.sheetnames)}")
        
        # Выбираем лист по индексу (учитывая, что индексация начинается с 0)
        sheet_name = wb.sheetnames[sheet_index - 1]
        ws = wb[sheet_name]
        logger.info(f"Выбран лист: {sheet_name}")
        
        # Определяем общее количество строк с артикулами
        total_rows = 0
        articles_list = []
        
        # Более тщательная проверка наличия артикулов
        for row in range(1, ws.max_row + 1):
            article_cell = f"{article_column}{row}"
            
            # Проверяем, что ячейка не пустая и содержит значение
            if ws[article_cell].value is not None:
                # Преобразуем значение в строку и проверяем, что оно не пустое после удаления пробелов
                article_value = str(ws[article_cell].value).strip()
                if article_value:
                    total_rows += 1
                    articles_list.append(article_value)
                    logger.debug(f"Найден артикул в строке {row}: {article_value}")
        
        stats["total_articles"] = total_rows
        logger.info(f"Найдено {total_rows} строк с артикулами")
        
        # Если нет артикулов, выводим подробную информацию и выходим
        if total_rows == 0:
            logger.warning(f"В файле не найдено артикулов в столбце {article_column}, начиная со строки {start_row}")
            logger.info(f"Проверьте следующее:")
            logger.info(f"1. Правильно ли указан столбец с артикулами ({article_column})?")
            logger.info(f"2. Правильно ли указана начальная строка ({start_row})?")
            logger.info(f"3. Содержит ли выбранный лист ({sheet_name}) данные?")
            
            # Сохраняем пустой результат
            output_file = f"{os.path.splitext(excel_file)[0]}_with_images.xlsx"
            wb.save(output_file)
            stats["end_time"] = time.time()
            stats["processing_time"] = stats["end_time"] - stats["start_time"]
            stats["output_file"] = output_file
            
            return stats
        
        # Если нужно настроить размер ячейки, делаем это сразу
        if adjust_cell_size:
            # Устанавливаем ширину колонки (1 единица ≈ 0.1640625 символа)
            col_width = column_width / 7  # Приблизительное преобразование пикселей в единицы ширины столбца
            ws.column_dimensions[image_column].width = col_width
            logger.info(f"Установлена ширина колонки {image_column}: {col_width} ед. ({column_width} пикс.)")
        
        # Обрабатываем каждую строку
        logger.info("Начинаем обработку строк...")
        for row in range(1, ws.max_row + 1):
            article_cell = f"{article_column}{row}"
            article_value = ws[article_cell].value
            
            # Пропускаем пустые ячейки
            if article_value is None:
                continue
                
            # Преобразуем значение в строку и проверяем, что оно не пустое после удаления пробелов
            article = str(article_value).strip()
            if not article:
                continue
            
            logger.debug(f"Обрабатываем артикул: {article} (строка {row})")
            
            # Ищем изображение по артикулу
            image_path = image_utils.find_image_by_article(article, images_folder)
            
            if image_path:
                stats["images_found"] += 1
                logger.info(f"Найдено изображение для артикула '{article}': {image_path}")
                
                # Обрабатываем изображение
                try:
                    # Получаем исходные размеры изображения
                    original_width, original_height = image_utils.get_image_dimensions(image_path)
                    logger.debug(f"Оригинальные размеры изображения: {original_width}x{original_height}")
                    
                    # Обрабатываем изображение только с оптимизацией качества,
                    # но без принудительного изменения размеров
                    img_buffer = image_utils.optimize_image_for_excel(
                        image_path=image_path,
                        max_size_kb=max_size_kb
                    )
                    
                    logger.debug(f"Изображение оптимизировано для вставки в Excel (размер файла не более {max_size_kb}KB)")
                    
                    # Вставляем изображение в ячейку
                    try:
                        # Получаем букву столбца и номер строки
                        cell_address = f"{image_column}{row}"
                        
                        # Если нужно настроить высоту строки
                        if adjust_cell_size:
                            # Устанавливаем высоту строки (1 единица ≈ 0.75 пункта)
                            row_height_excel = row_height * 0.75  # Приблизительное преобразование пикселей в единицы высоты строки
                            ws.row_dimensions[row].height = row_height_excel
                            logger.debug(f"Установлена высота строки {row}: {row_height_excel} ед. ({row_height} пикс.)")
                        
                        # Определяем ширину и высоту на основе пропорций изображения
                        aspect_ratio = original_height / original_width if original_width > 0 else 1.0
                        width_px = row_height  # Используем высоту строки как базовую ширину
                        height_px = int(width_px * aspect_ratio)
                        
                        # Используем insert_image_from_buffer вместо создания временного файла
                        img_buffer.seek(0)  # Сбрасываем указатель в начало буфера
                        insert_image_from_buffer(
                            worksheet=ws,
                            image_buffer=img_buffer,
                            anchor_cell=cell_address,
                            width=width_px,
                            height=height_px,
                            preserve_aspect_ratio=True,
                            background_color="000000"  # Добавляем черный фон
                        )
                        
                        stats["images_inserted"] += 1
                        logger.info(f"Изображение вставлено в ячейку {cell_address}")
                        
                    except Exception as e:
                        logger.error(f"Ошибка при вставке изображения в ячейку {cell_address}: {e}")
                        # Продолжаем обработку других строк
                
                except Exception as e:
                    logger.error(f"Ошибка при обработке изображения для артикула '{article}': {e}")
                    # Продолжаем обработку других строк
            else:
                logger.warning(f"Изображение для артикула '{article}' не найдено")
        
        # Сохраняем файл
        output_file = f"{os.path.splitext(excel_file)[0]}_with_images.xlsx"
        logger.info(f"Сохраняем результат в файл: {output_file}")
        
        # Если файл уже существует, сначала удаляем его
        if os.path.exists(output_file):
            os.remove(output_file)
            logger.info(f"Удален существующий файл: {output_file}")
        
        wb.save(output_file)
        logger.info(f"Файл сохранен: {output_file}")
        
        # Обновляем статистику
        stats["end_time"] = time.time()
        stats["processing_time"] = stats["end_time"] - stats["start_time"]
        stats["output_file"] = output_file
        
        logger.info(f"Обработка завершена. Статистика: артикулов - {stats['total_articles']}, " + 
                    f"найдено изображений - {stats['images_found']}, вставлено - {stats['images_inserted']}")
        
        return stats
        
    except Exception as e:
        logger.exception(f"Ошибка при обработке Excel файла: {e}")
        raise 

def column_letter_to_index(column_letter: str) -> int:
    """
    Преобразует буквенное обозначение столбца (A, B, C, AA, AB, etc.) или числовое (1, 2, 3...) в индекс (0-based).
    
    Args:
        column_letter (str): Буквенное или числовое обозначение столбца
        
    Returns:
        int: Индекс столбца (0-based)
        
    Raises:
        ValueError: Если обозначение столбца неверное
    """
    try:
        # Если входное значение - число
        if column_letter.isdigit():
            col_idx = int(column_letter) - 1  # Преобразуем в 0-based индекс
            logger.debug(f"Преобразование числового обозначения {column_letter} в индекс {col_idx}")
            return col_idx
            
        # Если входное значение - буква
        col_idx = column_index_from_string(column_letter) - 1
        logger.debug(f"Преобразование буквенного обозначения {column_letter} в индекс {col_idx}")
        return col_idx
    except Exception as e:
        logger.error(f"Ошибка при преобразовании обозначения столбца '{column_letter}' в индекс: {e}")
        raise ValueError(f"Неверное обозначение колонки: '{column_letter}'. Ошибка: {e}")

def insert_images_to_excel(writer, df, image_column):
    """
    Вставляет изображения в файл Excel на основе данных из DataFrame
    
    Args:
        writer: ExcelWriter объект для записи в Excel
        df: DataFrame с путями к изображениям
        image_column: Название колонки с путями к изображениям
        
    Returns:
        bool: True, если успешно
    """
    try:
        if image_column is None or image_column not in df.columns:
            logger.warning(f"Колонка {image_column} не найдена в DataFrame")
            return False
            
        # Получаем рабочий лист
        worksheet = writer.sheets[list(writer.sheets.keys())[0]]
        
        # Находим индекс колонки с изображениями
        col_idx = list(df.columns).index(image_column) + 1  # +1 для совместимости с openpyxl (индексы с 1)
        col_letter = get_column_letter(col_idx)
        
        # Добавляем изображения в каждую ячейку
        row_offset = 1  # Учитываем строку заголовка
        
        for idx, row in df.iterrows():
            if pd.notna(row[image_column]):
                image_paths = str(row[image_column]).split(",")
                
                for i, img_path in enumerate(image_paths):
                    if os.path.exists(img_path.strip()):
                        # Определяем ячейку для вставки
                        anchor_cell = f"{col_letter}{idx + row_offset + 1}"
                        
                        # Вставляем изображение с черным фоном
                        insert_image(worksheet, img_path.strip(), anchor_cell, background_color="000000")
                        
        logger.info(f"Изображения успешно вставлены в Excel файл")
        return True
    except Exception as e:
        logger.error(f"Ошибка при вставке изображений в Excel: {e}")
        return False

import utils.image_utils as image_utils

def save_dataframe_with_images(excel_file: str, df: pd.DataFrame,
                         article_column: str, image_column: str, images_folder: str,
                         max_size_kb: int = 100,
                         adjust_cell_size: bool = True,
                         row_height: int = 120,
                         find_images_recursive: bool = True) -> Dict[str, Any]:
    image_utils.cached_quality = None # Reset cached quality for this processing session
    """
    Сохраняет DataFrame в Excel файл с изображениями на основе артикулов
    
    Args:
        excel_file (str): Путь к Excel файлу для сохранения
        df (pd.DataFrame): DataFrame с данными
        article_column (str): Название колонки с артикулами
        image_column (str): Название колонки для вставки изображений
        images_folder (str): Путь к папке с изображениями
        max_size_kb (int): Максимальный размер файла изображения в KB
        adjust_cell_size (bool): Настраивать ли размер ячейки автоматически
        row_height (int): Высота строки в пикселях (если adjust_cell_size=True)
        find_images_recursive (bool): Искать ли изображения рекурсивно в подпапках
        
    Returns:
        Dict[str, Any]: Статистика обработки
    """
    try:
        logger.info(f"Начинаем обработку Excel файла: {excel_file}")
        logger.info(f"Артикулы в колонке: {article_column}, изображения в колонке: {image_column}")
        logger.info(f"Папка с изображениями: {images_folder}")
        
        # Проверяем входные данные
        if not os.path.isdir(images_folder):
            logger.error(f"Папка с изображениями не найдена: {images_folder}")
            raise FileNotFoundError(f"Папка с изображениями не найдена: {images_folder}")
            
        if article_column not in df.columns:
            logger.error(f"Колонка с артикулами '{article_column}' не найдена в DataFrame")
            raise ValueError(f"Колонка с артикулами '{article_column}' не найдена в DataFrame")
            
        # Статистика
        stats = {
            "start_time": time.time(),
            "total_articles": 0,
            "images_found": 0,
            "images_inserted": 0,
            "processing_time": 0,
            "output_file": ""
        }
        
        # Создаем рабочую книгу Excel и лист
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Проставляем заголовки
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx).value = col_name
        
        # Устанавливаем ширину колонки с изображениями
        # Находим индекс колонки с изображениями
        image_column_index = None
        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name == image_column:
                image_column_index = col_idx
                break
        
        if image_column_index is not None:
            image_column_letter = get_column_letter(image_column_index)
            ws.column_dimensions[image_column_letter].width = 20  # Примерная ширина для изображений
            logger.debug(f"Установлена ширина колонки с изображениями {image_column_letter}: 20")
        else:
            logger.warning(f"Колонка с изображениями '{image_column}' не найдена в DataFrame")
            image_column_letter = 'B'  # Значение по умолчанию
        
        # Заполняем данные и вставляем изображения
        for df_idx, row_data in df.iterrows():
            excel_row = df_idx + 2  # +2 из-за заголовка и 1-индексации Excel
            
            # Заполняем данные из DataFrame
            for col_idx, col_name in enumerate(df.columns, 1):
                ws.cell(row=excel_row, column=col_idx).value = row_data[col_name]
            
            # Обрабатываем изображение
            article = row_data.get(article_column)
            if pd.notna(article) and article:
                stats["total_articles"] += 1
                
                # Ищем изображение по артикулу
                found_images = image_utils.find_images_by_article_name(
                    article, 
                    images_folder,
                    search_recursively=find_images_recursive
                )
                
                if found_images:
                    stats["images_found"] += 1
                    image_path = found_images[0]  # Берем первое найденное изображение
                    logger.info(f"Найдено изображение для артикула '{article}': {image_path}")
                    
                    try:
                        # Получаем оригинальные размеры изображения
                        original_width, original_height = image_utils.get_image_dimensions(image_path)
                        logger.debug(f"Оригинальные размеры изображения: {original_width}x{original_height}")
                        
                        # Обрабатываем изображение только с оптимизацией качества,
                        # но без принудительного изменения размеров
                        img_buffer = image_utils.optimize_image_for_excel(
                            image_path=image_path,
                            max_size_kb=max_size_kb
                        )
                        
                        logger.debug(f"Изображение оптимизировано для вставки в Excel (размер файла не более {max_size_kb}KB)")
                        
                        # Вставляем изображение в ячейку
                        try:
                            # Получаем букву столбца и номер строки
                            cell_address = f"{image_column_letter}{excel_row}"
                            
                            # Если нужно настроить высоту строки
                            if adjust_cell_size:
                                # Устанавливаем высоту строки (1 единица ≈ 0.75 пункта)
                                row_height_excel = row_height * 0.75  # Приблизительное преобразование пикселей в единицы высоты строки
                                ws.row_dimensions[excel_row].height = row_height_excel
                                logger.debug(f"Установлена высота строки {excel_row}: {row_height_excel} ед. ({row_height} пикс.)")
                            
                            # Определяем ширину и высоту на основе пропорций изображения
                            aspect_ratio = original_height / original_width if original_width > 0 else 1.0
                            width_px = row_height  # Используем высоту строки как базовую ширину
                            height_px = int(width_px * aspect_ratio)
                            
                            # Используем insert_image_from_buffer вместо создания временного файла
                            img_buffer.seek(0)  # Сбрасываем указатель в начало буфера
                            insert_image_from_buffer(
                                worksheet=ws,
                                image_buffer=img_buffer,
                                anchor_cell=cell_address,
                                width=width_px,
                                height=height_px,
                                preserve_aspect_ratio=True,
                                background_color="000000"  # Добавляем черный фон
                            )
                            
                            stats["images_inserted"] += 1
                            logger.info(f"Изображение вставлено в ячейку {cell_address}")
                            
                        except Exception as e:
                            logger.error(f"Ошибка при вставке изображения в ячейку {cell_address}: {e}")
                            # Продолжаем обработку других строк
                    
                    except Exception as e:
                        logger.error(f"Ошибка при обработке изображения для артикула '{article}': {e}")
                        # Продолжаем обработку других строк
                else:
                    logger.warning(f"Изображение для артикула '{article}' не найдено")
            else:
                logger.warning(f"Изображение для артикула '{article}' не найдено")
        
        # Сохраняем файл
        output_file = f"{os.path.splitext(excel_file)[0]}_with_images.xlsx"
        logger.info(f"Сохраняем результат в файл: {output_file}")
        
        # Если файл уже существует, сначала удаляем его
        if os.path.exists(output_file):
            os.remove(output_file)
            logger.info(f"Удален существующий файл: {output_file}")
        
        wb.save(output_file)
        logger.info(f"Файл сохранен: {output_file}")
        
        # Обновляем статистику
        stats["end_time"] = time.time()
        stats["processing_time"] = stats["end_time"] - stats["start_time"]
        stats["output_file"] = output_file
        
        logger.info(f"Обработка завершена. Статистика: артикулов - {stats['total_articles']}, " + 
                    f"найдено изображений - {stats['images_found']}, вставлено - {stats['images_inserted']}")
        
        return stats
        
    except Exception as e:
        logger.exception(f"Ошибка при обработке Excel файла: {e}")
        raise

def get_column_width_pixels(worksheet: Worksheet, column_letter: str) -> int:
    """
    Получает ширину колонки в пикселях на основе настроек Excel.
    
    Args:
        worksheet (Worksheet): Рабочий лист
        column_letter (str): Буква колонки (например, 'A', 'B', etc.)
        
    Returns:
        int: Ширина колонки в пикселях
    """
    try:
        # Получаем размер колонки из объекта column_dimensions
        width_in_excel_units = worksheet.column_dimensions[column_letter].width
        
        # Если width не установлен, используем значение по умолчанию
        if width_in_excel_units is None:
            width_in_excel_units = 8.43  # Стандартная ширина колонки в Excel
            
        # Более точный коэффициент преобразования единиц Excel в пиксели (около 7 пикселей на единицу)
        pixels = int(width_in_excel_units * 7.0)
        
        # Применяем небольшую корректировку для более точного соответствия
        pixels = int(pixels * 0.95)  # Уменьшаем на 5% для более точного соответствия
        
        logger.debug(f"Ширина колонки {column_letter}: {width_in_excel_units} ед. Excel ≈ {pixels} пикселей")
        return pixels
    except Exception as e:
        logger.error(f"Ошибка при получении ширины колонки {column_letter}: {e}")
        return 300  # Значение по умолчанию в пикселях

def set_cell_background(worksheet: Worksheet, cell_reference: str, color: str = "000000") -> bool:
    """
    Устанавливает цвет заливки ячейки.
    
    Args:
        worksheet (Worksheet): Рабочий лист
        cell_reference (str): Ссылка на ячейку (например, 'A1')
        color (str, optional): Цвет заливки в формате RRGGBB. По умолчанию "000000" (черный).
    
    Returns:
        bool: True, если успешно
    """
    try:
        # Получаем букву столбца и номер строки
        column_letter = ''.join(filter(str.isalpha, cell_reference))
        row_num = int(''.join(filter(str.isdigit, cell_reference)))
        
        # Получаем объект ячейки
        cell = worksheet.cell(row=row_num, column=column_index_from_string(column_letter))
        
        # Применяем заливку
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        
        logger.debug(f"Установлен цвет заливки ячейки {cell_reference}: {color}")
        return True
    except Exception as e:
        logger.error(f"Ошибка при установке цвета заливки ячейки {cell_reference}: {e}")
        return False